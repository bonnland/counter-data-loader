import openpyxl
import collections
import os


class TitleMasterReport:
    """
    Represents a Title Master Report.

    This class is able to load any of the standard views defined for the
    Title Master Report. Currently, only J1/J3 reports are loaded for
    journal titles and the B1/B3 reports for books. The type of standard view
    is specified in the header (Report_ID).
    """

    PERIODS = ['', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug',
        'sep', 'oct', 'nov', 'dec']
    MAX_COLS = 16384
    MAX_ROWS = 1048576
    HEADER_ROW = 14
    DATA_ROW_START = 15
    DATA_COL_START = 1

    def __init__(self, workbook):
        self._workbook = openpyxl.load_workbook(filename=workbook, data_only=True)
        self._worksheet = self._workbook.active
        self._filename = os.path.basename(workbook)
        self._report_id = self._worksheet.cell(row=2, column=2).value
        self._reporting_period = self._worksheet.cell(row=10, column=2).value

    @property
    def filename(self):
        return self._filename

    @property
    def report_id(self):
        return self._report_id

    @property
    def begin_date(self):
        kv_pair = self._reporting_period.split(';')[0]
        return kv_pair.split('=')[1]

    @property
    def end_date(self):
        kv_pair = self._reporting_period.split(';')[1]
        return kv_pair.split('=')[1]

    @property
    def title_type(self):
        return self.report_id[3:4]

    def _header_row(self):
        """
        Returns the report header row (column names).

        The header row is constructed from two pieces:

            - all the title identifying information
            - the usage date columns

        The exact number of columns depends on the report (view) type. Book
        reports include an ISBN and YOP while journal reports do not*. Also,
        Access_Type may be included in either report type, e.g., in TR_B3 and
        TR_J3.

        For the date columns, the exact number (and range) depends on the
        reporting period of the report. In most cases, the reporting period
        will be January to December; however, the report could have been
        run for any period, e.g., April to July. Therefore, the begin and
        end dates for the reporting period will determine which month columns
        to include in the header.

        *The one exception is the TR_J4 report, which also has a YOP column.
        However, there are no current plans to load this type of report so
        this exception can be ignored.
        """

        # Build the title/book info columns by iterating through the spreadsheet
        # columns beginning with A (Title)
        col = {'TR_J1': 11, 'TR_J3': 12, 'TR_B1': 13, 'TR_B3': 14}
        header = list()
        for row in self._worksheet.iter_cols(min_row=self.HEADER_ROW, min_col=1,
            max_row=self.HEADER_ROW, max_col=col[self.report_id], values_only=True):
            for cell in row:
                header.append(cell.lower())

        # Append the date columns
        start_month = int(self.begin_date[5:7])
        end_month = int(self.end_date[5:7])
        period_cols = self.PERIODS[start_month:end_month+1]
        for month in period_cols:
            header.append(month)

        return header

    def data_rows(self):
        """
        Returns the range of data rows. For all report types, this is rows
        15 and onwards. The actual number of rows in a given report is
        variable and depends on the publisher.
        """

        r = self.DATA_ROW_START
        for row in self._worksheet.iter_rows(min_row=self.DATA_ROW_START, min_col=1,
            max_row=self.MAX_ROWS, max_col=1, values_only=True):
            if row[0] is None: # When no more data, the first cell in the row will be blank
                break
            r += 1

        return list(range(self.DATA_ROW_START,r))

    def data_cols(self):
        """
        Returns the range of data columns. The actual number of columns
        depends on the report type.
        """

        c = 1
        for col in self._worksheet.iter_cols(min_row=self.HEADER_ROW,
            min_col=self.DATA_COL_START, max_row=self.HEADER_ROW,
            max_col=self.MAX_COLS, values_only=True):
            if col[0] is None:
                break
            c += 1

        return list(range(self.DATA_COL_START,c))

    def get_row(self, n):
        """
        Gets a data row from the spreadsheet for a given row number.

        The return value is a named tuple, which provides for more
        intuitive referencing of columns during inserts and updates.
        """

        row_spec = collections.namedtuple('ReportRow', self._header_row())

        datarow = list()
        for row in self._worksheet.iter_cols(min_row=n, min_col=1, max_row=n,
            max_col=max(self.data_cols()), values_only=True):
            for cell in row:
                if cell is None: # Check for empty cells
                    datarow.append('')
                else:
                    datarow.append(str(cell).replace('"', '')) # Remove quotes

        return row_spec._make(datarow)